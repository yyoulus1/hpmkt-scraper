"""
scraper.py
----------
Playwright-based scraper for the High Point Market exhibitor directory.

Reads:  urls.xlsx  (one exhibitor URL per row, column A)
Writes: exhibitors_output.xlsx  (one row per exhibitor, all fields as columns)

Fields extracted (based on confirmed page HTML):
  - Company Name          h1 inside .exhibitor-contain
  - Exhibitor ID          numeric ID from URL path
  - Location / Booth      first span inside .info-block p  (e.g. "IHFC - C206, Commerce, Floor 2")
  - Shuttle Stop          second span inside .info-block p
  - Neighborhood          third span inside .info-block p
  - Phone                 span containing "Corporate Phone:" text
  - Website               external link inside .info-block
  - Instagram             .info-block ul.social a.inst
  - Facebook              .info-block ul.social a.fb
  - YouTube (channel)     .info-block ul.social a.yt
  - Pinterest             .info-block ul.social a.pint
  - Twitter               .info-block ul.social a.twt  (if present)
  - LinkedIn              .info-block ul.social a.li   (if present)
  - Description           #whoweare tab content
  - Video URLs            YouTube iframes in #video tab (comma-separated)
  - Gallery Image URLs    img tags in #photos tab (comma-separated)
  - Source URL            the original input URL
  - Status                ok / timeout / error:<message>

Usage:
    TEST_MODE=3 python scraper.py     # scrape first 3 URLs only
    python scraper.py                 # scrape all URLs (full run)
"""

import asyncio
import json
import os
import re
import time
from pathlib import Path

from playwright.async_api import async_playwright, TimeoutError as PWTimeoutError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Configuration ────────────────────────────────────────────────────────────
INPUT_FILE    = "urls.xlsx"
OUTPUT_FILE   = "exhibitors_output.xlsx"
PROGRESS_FILE = "progress.json"

CONCURRENCY   = 3
RETRY_LIMIT   = 2
PAGE_TIMEOUT  = 30_000   # ms

# TEST_MODE: if set to a positive integer, only process that many URLs.
# Set via env var: TEST_MODE=3 python scraper.py
_test_env = os.environ.get("TEST_MODE", "0").strip()
TEST_MODE = int(_test_env) if _test_env.isdigit() else 0

BASE_URL = "https://www.highpointmarket.org"

# ── Excel column definitions ─────────────────────────────────────────────────
COLUMNS = [
    "Company Name",
    "Exhibitor ID",
    "Location / Booth",
    "Shuttle Stop",
    "Neighborhood",
    "Phone",
    "Website",
    "Instagram",
    "Facebook",
    "YouTube",
    "Pinterest",
    "Twitter",
    "LinkedIn",
    "Description",
    "Video URLs",
    "Gallery Image URLs",
    "Source URL",
    "Status",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def load_urls() -> list[str]:
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        val = row[0]
        if val and str(val).strip().startswith("http"):
            urls.append(str(val).strip())
    # Also handle sheets where row 1 is already a URL (no header)
    first = ws.cell(1, 1).value
    if first and str(first).strip().startswith("http"):
        urls.insert(0, str(first).strip())
    return list(dict.fromkeys(urls))  # deduplicate, preserve order


def load_progress() -> dict:
    if Path(PROGRESS_FILE).exists():
        with open(PROGRESS_FILE) as f:
            return json.load(f)
    return {}


def save_progress(progress: dict):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def extract_id_from_url(url: str) -> str:
    m = re.search(r"/exhibitor/(\d+)", url)
    return m.group(1) if m else ""


def clean_text(t: str) -> str:
    return re.sub(r"\s+", " ", t).strip()


# ── Page scraping ─────────────────────────────────────────────────────────────

async def scrape_page(page, url: str) -> dict:
    """Scrape a single exhibitor page and return a dict of field→value."""
    row = {col: "" for col in COLUMNS}
    row["Source URL"] = url
    row["Exhibitor ID"] = extract_id_from_url(url)

    await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)

    # Wait for the main content to appear
    try:
        await page.wait_for_selector(".exhibitor-contain h1", timeout=PAGE_TIMEOUT)
    except PWTimeoutError:
        row["Status"] = "timeout: h1 not found"
        return row

    # ── Company Name ────────────────────────────────────────────────────────
    try:
        row["Company Name"] = clean_text(
            await page.locator(".exhibitor-contain h1").first.inner_text()
        )
    except Exception:
        pass

    # ── Location / Booth, Shuttle Stop, Neighborhood, Phone ─────────────────
    # The .info-block contains a <p> with multiple <span> children:
    #   span 1: location/booth  (e.g. "IHFC - C206, Commerce, Floor 2")
    #   span 2: Shuttle Stop: ...
    #   span 3: Neighborhood: ...
    #   span N: Corporate Phone: ...
    try:
        spans = await page.locator(".info-block p span").all()
        span_texts = []
        for sp in spans:
            t = clean_text(await sp.inner_text())
            if t:
                span_texts.append(t)

        for text in span_texts:
            tl = text.lower()
            if "shuttle stop" in tl:
                row["Shuttle Stop"] = re.sub(r"(?i)shuttle stop\s*:\s*", "", text).strip()
            elif "neighborhood" in tl:
                row["Neighborhood"] = re.sub(r"(?i)neighborhood\s*:\s*", "", text).strip()
            elif "corporate phone" in tl:
                row["Phone"] = re.sub(r"(?i)corporate phone\s*:\s*", "", text).strip()
            elif not row["Location / Booth"]:
                # First non-labelled span is the location/booth
                row["Location / Booth"] = text

    except Exception:
        pass

    # ── Website ──────────────────────────────────────────────────────────────
    # External link inside .info-block (not a social icon, not a tel: link)
    try:
        links = await page.locator(".info-block a[href]").all()
        for link in links:
            href = (await link.get_attribute("href") or "").strip()
            if href.startswith("http") and "highpointmarket.org" not in href:
                # Exclude social icon hrefs (they'll be grabbed separately)
                parent_ul = await link.evaluate("el => el.closest('ul.social') ? 'yes' : 'no'")
                if parent_ul == "no":
                    row["Website"] = href
                    break
    except Exception:
        pass

    # ── Social Links (scoped to .info-block ul.social ONLY) ─────────────────
    social_container = ".info-block ul.social"
    social_map = {
        "Instagram": f"{social_container} a.inst",
        "Facebook":  f"{social_container} a.fb",
        "YouTube":   f"{social_container} a.yt",
        "Pinterest": f"{social_container} a.pint",
        "Twitter":   f"{social_container} a.twt",
        "LinkedIn":  f"{social_container} a.li",
    }
    for field, selector in social_map.items():
        try:
            el = page.locator(selector).first
            if await el.count() > 0:
                row[field] = (await el.get_attribute("href") or "").strip()
        except Exception:
            pass

    # ── Description (Who We Are tab) ─────────────────────────────────────────
    try:
        desc_el = page.locator("#whoweare").first
        if await desc_el.count() > 0:
            row["Description"] = clean_text(await desc_el.inner_text())
    except Exception:
        pass

    # ── Videos (iframes in #video tab) ───────────────────────────────────────
    try:
        iframes = await page.locator("#video iframe[src]").all()
        video_urls = []
        for iframe in iframes:
            src = (await iframe.get_attribute("src") or "").strip()
            if src:
                video_urls.append(src)
        row["Video URLs"] = " | ".join(video_urls)
    except Exception:
        pass

    # ── Gallery Images (#photos tab) ─────────────────────────────────────────
    try:
        imgs = await page.locator("#photos img[src]").all()
        img_urls = []
        for img in imgs:
            src = (await img.get_attribute("src") or "").strip()
            if src:
                full = src if src.startswith("http") else f"{BASE_URL}{src}"
                img_urls.append(full)
        row["Gallery Image URLs"] = " | ".join(img_urls)
    except Exception:
        pass

    row["Status"] = "ok"
    return row


async def scrape_with_retry(context, url: str, worker_id: int) -> dict:
    page = await context.new_page()
    try:
        for attempt in range(1, RETRY_LIMIT + 2):
            try:
                result = await scrape_page(page, url)
                return result
            except PWTimeoutError:
                if attempt <= RETRY_LIMIT:
                    print(f"  [W{worker_id}] Timeout on attempt {attempt}, retrying: {url}")
                    await asyncio.sleep(2)
                else:
                    return {col: "" for col in COLUMNS} | {
                        "Source URL": url,
                        "Exhibitor ID": extract_id_from_url(url),
                        "Status": "timeout",
                    }
            except Exception as e:
                if attempt <= RETRY_LIMIT:
                    print(f"  [W{worker_id}] Error attempt {attempt}: {e}, retrying: {url}")
                    await asyncio.sleep(2)
                else:
                    return {col: "" for col in COLUMNS} | {
                        "Source URL": url,
                        "Exhibitor ID": extract_id_from_url(url),
                        "Status": f"error: {str(e)[:120]}",
                    }
    finally:
        await page.close()

    # Should never reach here
    return {col: "" for col in COLUMNS} | {"Source URL": url, "Status": "error: unknown"}


# ── Worker ────────────────────────────────────────────────────────────────────

async def worker(worker_id: int, queue: asyncio.Queue, results: list, progress: dict, browser):
    while True:
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break

        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            )
        )
        try:
            print(f"  [W{worker_id}] Scraping: {url}")
            row = await scrape_with_retry(context, url, worker_id)
            results.append(row)
            progress[url] = row["Status"]
            save_progress(progress)
            print(f"  [W{worker_id}] ✓ {row['Company Name']} — {row['Status']}")
        finally:
            await context.close()
            queue.task_done()


# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(results: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibitors"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, row_data in enumerate(results, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width (capped at 80)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 80)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"\nSaved {len(results)} rows → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=== High Point Market Exhibitor Scraper ===")

    if not Path(INPUT_FILE).exists():
        print(f"ERROR: {INPUT_FILE} not found. Run harvest_urls.py first.")
        return

    all_urls = load_urls()
    print(f"Loaded {len(all_urls)} URLs from {INPUT_FILE}")

    # Load resume progress (skip in test mode)
    if TEST_MODE:
        print(f"TEST MODE: processing first {TEST_MODE} URLs (ignoring progress.json)")
        urls_to_scrape = all_urls[:TEST_MODE]
        progress = {}
    else:
        progress = load_progress()
        already_done = {u for u, s in progress.items() if s == "ok"}
        urls_to_scrape = [u for u in all_urls if u not in already_done]
        skipped = len(all_urls) - len(urls_to_scrape)
        if skipped:
            print(f"Resuming: skipping {skipped} already-completed URLs")

    print(f"URLs to scrape: {len(urls_to_scrape)}")

    if not urls_to_scrape:
        print("Nothing to do — all URLs already completed.")
        return

    queue: asyncio.Queue = asyncio.Queue()
    for url in urls_to_scrape:
        queue.put_nowait(url)

    results: list[dict] = []
    concurrency = min(CONCURRENCY, len(urls_to_scrape))
    t0 = time.time()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        workers = [
            asyncio.create_task(worker(i + 1, queue, results, progress, browser))
            for i in range(concurrency)
        ]
        await asyncio.gather(*workers)
        await browser.close()

    elapsed = time.time() - t0
    ok     = sum(1 for r in results if r["Status"] == "ok")
    errors = len(results) - ok
    print(f"\nDone in {elapsed:.1f}s — {ok} ok, {errors} errors")

    write_excel(results, OUTPUT_FILE)

    if errors:
        print("\nFailed URLs:")
        for r in results:
            if r["Status"] != "ok":
                print(f"  {r['Status']:30s}  {r['Source URL']}")


if __name__ == "__main__":
    asyncio.run(main())
