"""
harvest_urls.py
---------------
Crawls the High Point Market exhibitor directory and collects all individual
exhibitor URLs into urls.xlsx.

Fixes applied vs v1:
  - Use 'domcontentloaded' instead of 'networkidle' — the page never fully
    settles due to analytics scripts, causing 30s timeouts on busy letters
  - Longer timeout (60s) for the initial page load
  - After page load, wait explicitly for the exhibitor links to appear
  - Aggressive "Load More" loop — keep clicking until the button disappears
  - Reduced concurrency to 2 to avoid rate-limiting / connection drops
  - Each letter gets its own fresh browser context to avoid shared-state issues
  - Retry logic: if a letter errors, try once more before giving up
"""

import asyncio
import re
import time
import urllib.parse

from playwright.async_api import async_playwright, TimeoutError as PWTimeoutError
import openpyxl

BASE_URL      = "https://www.highpointmarket.org"
DIRECTORY_URL = f"{BASE_URL}/ExhibitorDirectory"

ALPHA_LETTERS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["0-9"]

CONCURRENCY  = 2          # parallel letters — keep low to avoid rate limiting
OUTPUT_FILE  = "urls.xlsx"
LOAD_TIMEOUT = 60_000     # 60s for initial page load
SEL_TIMEOUT  = 20_000     # 20s waiting for exhibitor links to appear


async def get_urls_for_letter(browser, letter: str, attempt: int = 1) -> list[str]:
    """Open one directory page filtered by alpha letter and collect all exhibitor links."""
    context = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1280, "height": 900},
    )
    page = await context.new_page()

    try:
        # URL-encode the filter JSON — no spaces, clean encoding
        filter_json = f'{{"Type":"Alpha","Values":["{letter}"]}}'
        encoded = urllib.parse.quote(filter_json)
        url = f"{DIRECTORY_URL}?filters={encoded}"

        # domcontentloaded is much faster and reliable vs networkidle
        await page.goto(url, wait_until="domcontentloaded", timeout=LOAD_TIMEOUT)

        # Give JS a moment to kick off the API call that loads exhibitor cards
        await page.wait_for_timeout(3_000)

        # Wait for at least one exhibitor link to appear
        try:
            await page.wait_for_selector("a[href*='/exhibitor/']", timeout=SEL_TIMEOUT)
        except PWTimeoutError:
            print(f"  [{letter}] No exhibitors found (or page timed out waiting for cards)")
            return []

        # ── Load More loop ──────────────────────────────────────────────────
        load_more_sel = (
            "button:has-text('Load More'), "
            "a:has-text('Load More'), "
            "button:has-text('Show More'), "
            "a:has-text('Show More')"
        )

        click_count = 0
        while True:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1_500)

            load_more = page.locator(load_more_sel)
            if await load_more.count() == 0:
                break

            try:
                await load_more.first.scroll_into_view_if_needed(timeout=3_000)
                await load_more.first.click(timeout=5_000)
                click_count += 1
                print(f"  [{letter}] clicked Load More ({click_count}x)...")
                await page.wait_for_timeout(2_000)
            except Exception:
                break

        # ── Extract all /exhibitor/<id> links ───────────────────────────────
        hrefs = await page.locator("a[href*='/exhibitor/']").evaluate_all(
            "els => els.map(e => e.getAttribute('href'))"
        )
        urls = set()
        for href in hrefs:
            if href and re.match(r"^/exhibitor/\d+$", href):
                urls.add(f"{BASE_URL}{href}")

        suffix = f" (clicked Load More {click_count}x)" if click_count else ""
        print(f"  [{letter}] found {len(urls)} exhibitors{suffix}")
        return sorted(urls)

    except Exception as e:
        msg = str(e)[:120]
        if attempt < 3:
            print(f"  [{letter}] ERROR attempt {attempt}: {msg} — retrying in 5s...")
            await context.close()
            await asyncio.sleep(5)
            return await get_urls_for_letter(browser, letter, attempt + 1)
        else:
            print(f"  [{letter}] FAILED after {attempt} attempts: {msg}")
            return []

    finally:
        try:
            await context.close()
        except Exception:
            pass


async def harvest_all() -> list[str]:
    all_urls: set[str] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )

        semaphore = asyncio.Semaphore(CONCURRENCY)

        async def bounded(letter):
            async with semaphore:
                return await get_urls_for_letter(browser, letter)

        tasks = [bounded(letter) for letter in ALPHA_LETTERS]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        for result in results:
            if isinstance(result, list):
                all_urls.update(result)
            elif isinstance(result, Exception):
                print(f"  Unexpected exception: {result}")

        await browser.close()

    return sorted(all_urls)


def save_to_excel(urls: list[str], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibitor URLs"

    ws["A1"] = "URL"
    ws["A1"].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    ws["A1"].fill = openpyxl.styles.PatternFill("solid", fgColor="1F3864")

    for i, url in enumerate(urls, start=2):
        ws.cell(row=i, column=1, value=url)

    ws.column_dimensions["A"].width = 60
    wb.save(path)
    print(f"\nSaved {len(urls)} URLs → {path}")


async def main():
    print("=== High Point Market — URL Harvester ===")
    print(f"Crawling {len(ALPHA_LETTERS)} alpha buckets with concurrency={CONCURRENCY}...\n")
    t0 = time.time()

    urls = await harvest_all()

    elapsed = time.time() - t0
    print(f"\nTotal unique exhibitor URLs found: {len(urls)}")
    print(f"Elapsed: {elapsed:.1f}s")

    if urls:
        save_to_excel(urls, OUTPUT_FILE)
    else:
        print("WARNING: No URLs found — check if the site structure has changed.")


if __name__ == "__main__":
    asyncio.run(main())
